import tensorflow as tf
import numpy as np
import matplotlib.pyplot as plt
import tkinter
from tkinter import filedialog
import os
import openpyxl

def selectFile():
    root = tkinter.Tk()
    root.withdraw() #use to hide tkinter window
###########         ###########         ###########         ###########         ###########         
    def search_for_excel1 ():
        currdir = os.getcwd()
        tempdir = filedialog.askopenfilename(parent=root, initialdir=currdir, title='Please select a directory')
        if len(tempdir) > 0:
            print ("You chose: %s" % tempdir)
        return tempdir
    
    def search_for_excel2 ():
        currdir = os.getcwd()
        tempdir = filedialog.askopenfilename(parent=root, initialdir=currdir, title='Please select a directory')
        if len(tempdir) > 0:
            print ("You chose: %s" % tempdir)
        return tempdir
###########         ###########         ###########         ###########         ###########         
    def compare_excel_cells(file1_path, file2_path):
       
        # Load the first Excel file
        workbook1 = openpyxl.load_workbook(file1_path)
        sheet1 = workbook1.active

        # Load the second Excel file
        workbook2 = openpyxl.load_workbook(file2_path)
        sheet2 = workbook2.active

        # Get the dimensions of both sheets
        max_row = max(sheet1.max_row, sheet2.max_row)
        max_column = max(sheet1.max_column, sheet2.max_column)

        # Compare cell values
        for row in range(1, max_row + 1,1):
            for column in range(1, max_column + 1,1):
                cell1 = sheet1.cell(row=row, column=column)
                cell2 = sheet2.cell(row=row, column=column)
###########         ###########         ###########         ###########         ###########         
                if cell1.value != cell2.value:
                    with open("C:/Users/efsan/Documents/GitHub/Overlap/overlapMisMatch.txt", 'a') as file:                
                        file.write("\n")
                        file.write(f"Cell mismatch at Row {row}, Column {column}:\n")
                        file.write(f"File 1: {cell1.value}\n")
                        file.write(f"File 2: {cell2.value}\n")
                        file.write("")
                        file.write("\n")

                if cell1.value == cell2.value:
                    with open("C:/Users/efsan/Documents/GitHub/Overlap/overlapMatch.txt", 'a') as file:                
                        file.write("\n")
                        file.write(f"Cell match at Row {row}, Column {column}:\n")
                        file.write(f"{file1_path}: {cell1.value}\n")
                        file.write(f"{file2_path}: {cell2.value}\n")
                        file.write("")
                        file.write("\n")
###########         ###########         ###########         ###########         ###########         
        print("Comparison completed.")
###########################################################################################
    # Provide the paths of the Excel files to compare
    file1_path = search_for_excel1()
    print ("\nfile_path_variable = ", file1_path)
    file2_path = search_for_excel2()
    print ("\nfile_path_variable = ", file2_path)
###########         ###########         ###########         ###########         ###########         
    compare_excel_cells(file1_path, file2_path)
